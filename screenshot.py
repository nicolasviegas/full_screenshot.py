# -*- coding: utf-8 -*-
"""
Created on Thu Aug 25 14:33:34 2022

@author: nviegas001
"""

# Python program to read an excel file

import openpyxl
import time
import os
import shutil
import tkinter as tk
from tkinter import simpledialog
from tkinter import messagebox

from selenium import webdriver
from selenium.webdriver.common.by import By

import cv2
import numpy as np

from PIL import Image



# PATH ES LA DIRECCION DONDE SE ENCUENTRA EL EXCEL
#PARENT DIR ES LA DIR DONDE SE VA A EJECUTAR EL SCRIPT Y DONDE SE VAN A CREAR LAS CARPETAS DE LOS CASOS
#path = "C:\\Users\\nviegas001\\python-scripts\\test-excel.xlsx"
#parent_dir = "C:\\Users\\nviegas001\\python-scripts\\"
download_dir = "C:\\Users\\nviegas001\\Downloads\\"


ROOT = tk.Tk()
ROOT.withdraw()
excel = simpledialog.askstring(title="Evidencia", prompt="Ingresa la direccion del excel donde se encuentran las muestras(TIENE QUE TERMINAR CON .xlsx)")
path = excel

ROOT = tk.Tk()
ROOT.withdraw()
columna = simpledialog.askstring(title="Evidencia", prompt="Ingresa el nro de la columna en donde se encuentran los codigos en el excel")
nro_columna = int(columna)


ROOT = tk.Tk()
ROOT.withdraw()
script = simpledialog.askstring(title="Evidencia", prompt="Ingresa la direccion de la carpeta donde se guardara la evidencia")
parent_dir = script + '\\'

ROOT = tk.Tk()
ROOT.withdraw()
username = simpledialog.askstring(title="Usuario", prompt="Ingrese el nombre de usuario de Jira")

ROOT = tk.Tk()
ROOT.withdraw()
psJira = simpledialog.askstring(title="Password", prompt="Ingrese password de Jira")

# =============================================================================
# ROOT = tk.Tk()
# ROOT.withdraw()
# download = simpledialog.askstring(title="Evidencia", prompt="Ingresa la direccion de la carpeta de descargas(AGREGAR UNA \ AL FINAL DE LA DIR)")
# download_dir = download
# =============================================================================

url = 'https://jira.despegar.com/browse/'
url_word = 'https://jira.despegar.com/si/jira.issueviews:issue-word/'


wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

cell_obj = sheet_obj.cell(row = 1, column = 1)
max_col = sheet_obj.max_column
max_r = sheet_obj.max_row



#Si no hay AN100, NI HAY PASAJE A PRODUCCION, COMENTAR LAS LINEAS DE MKDIR
def crear_directorio_por_caso():
    for i in range(2, max_r + 1):
        if i-1 < 10:
            directory = "Caso 0" + str(i - 1)
            dir_path = os.path.join(parent_dir, directory)
            os.mkdir(dir_path)
            an100 = "AN100"
            path_ann = os.path.join(directory, an100 )
            #os.mkdir(path_ann)
            pasaje_a_prod = "Pasaje a producción"
            path_prod = os.path.join(directory, pasaje_a_prod )
           # os.mkdir(path_prod)
        else:
            directory = "Caso " + str(i - 1)
            dir_path = os.path.join(parent_dir, directory)
            os.mkdir(dir_path)
            an100 = "AN100"
            path_ann = os.path.join(directory, an100 )
           # os.mkdir(path_ann)
            pasaje_a_prod = "Pasaje a producción"
            path_prod = os.path.join(directory, pasaje_a_prod )
         #   os.mkdir(path_prod)
       
crear_directorio_por_caso()

def export_as_word(ticket):
    chrome_options = webdriver.ChromeOptions()
   # chrome_options.add_argument('--headless')
    chrome_options.add_argument("--start-maximized")
    #chromedriver = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'chromedriver.exe')
   
   
    chrome = webdriver.Chrome(executable_path=r"C:\Users\nviegas001\python-scripts\webdriver\chromedriver.exe", options=chrome_options)

    link_word_ticket = url_word + str(ticket) + "/" + str(ticket) + ".doc"

    chrome.get(link_word_ticket)
   
    time.sleep(7)

    chrome.find_element(By.ID, "login-form-username").send_keys(username)

    chrome.find_element(By.ID, "login-form-password").send_keys(psJira)

    chrome.find_element(By.ID, "login-form-submit").click()
   
    time.sleep(7)
   
    current = download_dir + str(jira_ticket_num) + ".doc"
   
    dest = dir_path
   
    shutil.move(current, dest)
   
    chrome.close()
   
    chrome.quit()


   
###################################################################################################################################################
#######3##esta funcion saca screen y lo mueve a la carpeta asociada###############3
def take_screenshot(link,ticket):
   
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument("--start-maximized")
    #chromedriver = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'chromedriver.exe')
       
   
    chrome = webdriver.Chrome(executable_path=r"C:\Users\nviegas001\python-scripts\webdriver\chromedriver.exe", options=chrome_options)

   
    chrome.maximize_window()


    chrome.get(link)

    time.sleep(2)

    chrome.find_element(By.ID, "login-form-username").send_keys(username)

    chrome.find_element(By.ID, "login-form-password").send_keys(psJira)

    chrome.find_element(By.ID, "login-form-submit").click()


    time.sleep(1)
   

    #########################

    chrome.find_element(By.ID, "all-tabpanel").click()


    time.sleep(1)

    chrome.find_element(By.ID, "key-val").click()


    #CLICKEA EN TODO, PERO HAY QUE HACER QUE EL CLICK VUELVA AL PRINCIPIO DE TODO

    total_height = chrome.execute_script("return document.body.parentNode.scrollHeight") + 50000


    chrome.set_window_size(1920, total_height)


    time.sleep(5)

    path = parent_dir + str(jira_ticket_num) + ".PNG"


    chrome.save_screenshot(path)


    ########################## ENCONTRAR COORDENADAS DE DONDE TERMINA LA IMAGEN/// AMARILLO: RGB: (255, 211,  81), VERDE: (20,137,44)

    im = cv2.imread(path)
    green = [44, 137, 20]

    Y, X = np.where(np.all(im==green,axis=2))

    print(X,Y)

    array_length = len(Y)
    last_element = Y[array_length - 1]

    #####

    im2 =  Image.open(path)
    #EL PRIMERO Y EL SEGUNDO TIENEN QUE SER 0, el tercero 1920,  el cuarto es la diferencia de el ultimo amarillo mas 2000 ejemplo
    im_crop = im2.crop((0, 0, 1920, last_element + 500))
    im_crop.save(path, quality=95)

    current = parent_dir + str(jira_ticket_num) + ".PNG"
    dest = dir_path
   
    shutil.move(current, dest)
   
    time.sleep(10)


    chrome.quit()
   
################################################################################################################################################    
#DONDE DICE COLUM=X (linea 148) HAY QUE CAMBIARLO POR EL NRO DE COLUMNA EN DONDE SE ENCUENTRA EL CODIGO JIRA
for i in range(2, max_r + 1):
    cell_obj = sheet_obj.cell(row = i, column = nro_columna)
   
    if i-1 < 10:
        directory = "Caso 0" + str(i - 1)
        dir_path = os.path.join(parent_dir, directory) + "\\"
    else:
        directory = "Caso " + str(i - 1)
        dir_path = os.path.join(parent_dir, directory) + "\\"
   
    jira_ticket_num = cell_obj.value

   
    link = url + str(jira_ticket_num)  
   
   
    take_screenshot(link,jira_ticket_num)
    #export_as_word(jira_ticket_num)
 

messagebox.showinfo("Evidencia", "Programa finalizado")   