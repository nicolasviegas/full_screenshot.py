# -*- coding: utf-8 -*-
"""
Created on Fri Mar 31 15:48:54 2023

@author: nviegas001
"""

# -*- coding: utf-8 -*-
"""
Created on Thu Aug 25 14:33:34 2022

@author: nviegas001
"""

# Python program to read an excel file

# import openpyxl module
import openpyxl
import time
import os
import shutil
#from pyautogui import *
import pyautogui
import pyperclip
#import keyboard
import tkinter as tk
from tkinter import simpledialog
from tkinter import messagebox

#from selenium import webdriver


# PATH ES LA DIRECCION DONDE SE ENCUENTRA EL EXCEL
#PARENT DIR ES LA DIR DONDE SE VA A EJECUTAR EL SCRIPT Y DONDE SE VAN A CREAR LAS CARPETAS DE LOS CASOS
#path = "C:\\Users\\nviegas001\\python-scripts\\test-excel.xlsx"
#parent_dir = "C:\\Users\\nviegas001\\python-scripts\\"
download_dir = "C:\\Users\\nviegas001\\Downloads\\"


ROOT = tk.Tk()
ROOT.withdraw()
excel = simpledialog.askstring(title="Evidencia", prompt="Ingresa la direccion del excel donde se encuentran las muestras(TIENE QUE TERMINAR CON .xlsx)")
path = excel

ROOT = tk.Tk()
ROOT.withdraw()
columna = simpledialog.askstring(title="Evidencia", prompt="Ingresa el nro de la columna en donde se encuentran los codigos en el excel")
nro_columna = int(columna)


ROOT = tk.Tk()
ROOT.withdraw()
script = simpledialog.askstring(title="Evidencia", prompt="Ingresa la direccion donde se corre el script (AGREGAR UNA \ AL FINAL DE LA DIR)")
parent_dir = script + '\\'

# =============================================================================
# ROOT = tk.Tk()
# ROOT.withdraw()
# download = simpledialog.askstring(title="Evidencia", prompt="Ingresa la direccion de la carpeta de descargas(AGREGAR UNA \ AL FINAL DE LA DIR)")
# download_dir = download
# =============================================================================

time.sleep(2)




wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

cell_obj = sheet_obj.cell(row = 1, column = 1)
max_col = sheet_obj.max_column
max_r = sheet_obj.max_row


def click_buscador_jira():
    pyautogui.click(x=1712, y=173)
    with pyautogui.hold('ctrl',):
        pyautogui.press('a')
        pyautogui.press('backspace')
    
    
    
def escribir_nro_ticket(ticket):
    pyautogui.write(ticket)

#Si no hay AN100, NI HAY PASAJE A PRODUCCION, COMENTAR LAS LINEAS DE MKDIR
def crear_directorio_por_caso():
    for i in range(2, max_r + 1):
        if i-1 < 10:
            directory = "Caso 0" + str(i - 1)
            dir_path = os.path.join(parent_dir, directory)
            os.mkdir(dir_path)  
            an100 = "AN100"
        
            path_ann = os.path.join(dir_path, an100 )
          #  os.mkdir(path_ann) # comentar
            pasaje_a_prod = "Pasaje a producción"
           
            path_prod = os.path.join(dir_path, pasaje_a_prod )
           # os.mkdir(path_prod) # comentar
        else:
            directory = "Caso " + str(i - 1)
            dir_path = os.path.join(parent_dir, directory)
            os.mkdir(dir_path)
            an100 = "AN100"
         
            path_ann = os.path.join(dir_path, an100 )
            #os.mkdir(path_ann)# comentar
            pasaje_a_prod = "Pasaje a producción"
         
            path_prod = os.path.join(dir_path, pasaje_a_prod )
            #os.mkdir(path_prod)# comentar
        
#crear_directorio_por_caso()

def pegar_valor_en_buscador(valor):
    click_buscador_jira()
    escribir_nro_ticket(valor)
    pyautogui.press('enter')
    time.sleep(10) #estaba en 3
    pyautogui.click(x=1794, y=492)
    time.sleep(1)

#    print(valor)

#######3##esta funcion saca screen y lo mueve a la carpeta asociada###############3
def take_screenshot(valor,iterador):
    time.sleep(1)
    
    myScreenshot = pyautogui.screenshot()
    myScreenshot.save(str(jira_ticket_num) + "-" + str(iterador) +  ".PNG")
  
    current = parent_dir + (str(jira_ticket_num) + "-" + str(iterador) +  ".PNG")
    dest = dir_path
    
    shutil.move(current, dest)
    
    print(current)
    print(iterador)
    
    time.sleep(2)
    
def copiar_link_en_txt(links):
    with pyautogui.hold('ctrl',):
        pyautogui.press('l')
        time.sleep(1)
        pyautogui.press('c')
    
    time.sleep(1)
    spam = pyperclip.paste()
    links.write(spam + "\n")
   # print(spam)
   
def export_as_word():
    pyautogui.press('home')
    time.sleep(1)
    pyautogui.click(x=1822, y=322)
    time.sleep(1)
    pyautogui.click(x=1800, y=383)
    
    time.sleep(7)
    current = download_dir + str(jira_ticket_num) + ".doc"
    dest = dir_path
    shutil.move(current, dest)
    pyautogui.click(x=1891, y=974)


    
    #DONDE DICE COLUM=X (linea 148) HAY QUE CAMBIARLO POR EL NRO DE COLUMNA EN DONDE SE ENCUENTRA EL CODIGO JIRA
for i in range(2, max_r + 1):
    cell_obj = sheet_obj.cell(row = i, column = nro_columna)
   
    if i-1 < 10:
        directory = "Caso 0" + str(i - 1)
        dir_path = os.path.join(parent_dir, directory) + "\\"
    else:
        directory = "Caso " + str(i - 1)
        dir_path = os.path.join(parent_dir, directory) + "\\"
    
    jira_ticket_num = cell_obj.value
    pyautogui.press('home')
    #time.sleep(0.5)
    #pegar_valor_en_buscador(jira_ticket_num)
    #time.sleep(0.5)
    iterator = 1
    
# =============================================================================
    while(not pyautogui.pixelMatchesColor(102, 975, (223, 226, 231),tolerance=20)):
        if(pyautogui.pixelMatchesColor(1906, 1006, (80, 80, 80),tolerance=10)):
           time.sleep(1)
           take_screenshot(jira_ticket_num,iterator)
           print("Entre por el if")
           time.sleep(0.5)
           pyautogui.press('pagedown')
           iterator += 1
        else:
           time.sleep(1)
           take_screenshot(jira_ticket_num,iterator)
           print("Entre por el else")
           break
# =============================================================================
        
    print("Termine de sacar screenshots")
    time.sleep(1)
#    export_as_word()
    
    
messagebox.showinfo("Evidencia", "Programa finalizado")   
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
################################  TESTING ZONE. ESTO ES PARA EL SCREEN WEB 
   # copiar_link_en_txt(links)
   # links.close()
   # links = open("links.txt","r")
   # url = links.readline()
   # print(url)
    
    
#    driver = webdriver.Chrome()
#    driver.get("https://jira.despegar.com/browse/FST-14434")
#    time.sleep(1)
#    driver.get_screenshot_as_file(str(jira_ticket_num) + ".PNG")

#    current = parent_dir + str(jira_ticket_num) + ".PNG"
#    dest = dir_path
  
#    shutil.move(current, dest)
      
#    time.sleep(2)
    
    ############################## zona de testeo



